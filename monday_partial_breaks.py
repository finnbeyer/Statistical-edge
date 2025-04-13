import pandas as pd
import numpy as np
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import logging
from typing import Dict, List, Any, Optional
from datetime import datetime

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

def load_data(file_path: str) -> pd.DataFrame:
    """
    Load and prepare the data for analysis.
    
    Args:
        file_path: Path to the CSV file containing candle data
        
    Returns:
        DataFrame with prepared data
    """
    try:
        df = pd.read_csv(file_path, sep=';')
        df['Date'] = pd.to_datetime(df['Date'])
        df['DayOfWeek'] = df['Date'].dt.dayofweek
        df['Week'] = df['Date'].dt.isocalendar().week
        df['Year'] = df['Date'].dt.isocalendar().year
        return df
    except Exception as e:
        logger.error(f"Error loading data: {str(e)}")
        raise

def analyze_partial_breaks(df: pd.DataFrame) -> tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]], int]:
    """
    Analyze Monday ranges and track partial breaks.
    
    Args:
        df: DataFrame containing the candle data
        
    Returns:
        Tuple containing:
        - only_high_broken: List of weeks where only high was broken
        - only_low_broken: List of weeks where only low was broken
        - neither_broken: List of weeks where neither was broken
        - both_broken: List of weeks where both were broken
        - total_mondays: Total number of Mondays analyzed
    """
    only_high_broken = []
    only_low_broken = []
    neither_broken = []
    both_broken = []
    total_mondays = 0

    for (week, year), week_data in df.groupby(['Week', 'Year']):
        monday_data = week_data[week_data['DayOfWeek'] == 0]
        if len(monday_data) == 0:
            continue
            
        total_mondays += 1
        monday_high = monday_data['High'].iloc[0]
        monday_low = monday_data['Low'].iloc[0]
        monday_date = monday_data['Date'].iloc[0]
        
        rest_of_week = week_data[week_data['DayOfWeek'].isin([1,2,3,4])]
        
        # Check high break
        high_break_days = rest_of_week[rest_of_week['High'] > monday_high]
        high_broken = not high_break_days.empty
        
        # Check low break
        low_break_days = rest_of_week[rest_of_week['Low'] < monday_low]
        low_broken = not low_break_days.empty
        
        # Store week data
        week_info = {
            'Date': monday_date,
            'Week': week,
            'Year': year,
            'Monday High': monday_high,
            'Monday Low': monday_low,
            'High Break Day': high_break_days['DayOfWeek'].iloc[0] if high_broken else None,
            'Low Break Day': low_break_days['DayOfWeek'].iloc[0] if low_broken else None
        }
        
        if high_broken and not low_broken:
            only_high_broken.append(week_info)
        elif low_broken and not high_broken:
            only_low_broken.append(week_info)
        elif not high_broken and not low_broken:
            neither_broken.append(week_info)
        else:
            both_broken.append(week_info)

    return only_high_broken, only_low_broken, neither_broken, both_broken, total_mondays

def day_to_name(day: Optional[int]) -> str:
    """
    Convert day number to day name.
    
    Args:
        day: Day number (1=Tuesday, 2=Wednesday, etc.) or None
        
    Returns:
        Day name as string or "Not Broken" if None
    """
    if day is None:
        return "Not Broken"
    day_names = {1: 'Tuesday', 2: 'Wednesday', 3: 'Thursday', 4: 'Friday'}
    return day_names.get(day, str(day))

def write_section(ws: Any, title: str, data: List[Dict[str, Any]], start_row: int) -> int:
    """
    Write a section of data to the Excel worksheet.
    
    Args:
        ws: Excel worksheet object
        title: Section title
        data: List of data dictionaries
        start_row: Starting row number
        
    Returns:
        Next available row number
    """
    # Style definitions
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    subheader_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    
    # Write title
    ws[f'A{start_row}'] = title
    ws[f'A{start_row}'].font = header_font
    ws[f'A{start_row}'].fill = header_fill
    
    if not data:
        ws[f'A{start_row + 1}'] = "No instances found"
        return start_row + 2
    
    # Write headers
    headers = ['Date', 'Week', 'Year', 'Monday High', 'Monday Low', 'High Break Day', 'Low Break Day']
    for col, header in enumerate(headers, 1):
        ws.cell(row=start_row + 1, column=col, value=header)
        ws.cell(row=start_row + 1, column=col).font = header_font
        ws.cell(row=start_row + 1, column=col).fill = subheader_fill
    
    # Write data
    for row_idx, week_data in enumerate(data, start_row + 2):
        ws.cell(row=row_idx, column=1, value=week_data['Date'].strftime('%Y-%m-%d'))
        ws.cell(row=row_idx, column=2, value=week_data['Week'])
        ws.cell(row=row_idx, column=3, value=week_data['Year'])
        ws.cell(row=row_idx, column=4, value=week_data['Monday High'])
        ws.cell(row=row_idx, column=5, value=week_data['Monday Low'])
        ws.cell(row=row_idx, column=6, value=day_to_name(week_data['High Break Day']))
        ws.cell(row=row_idx, column=7, value=day_to_name(week_data['Low Break Day']))
    
    return row_idx + 2

def create_excel_report(
    only_high_broken: List[Dict[str, Any]],
    only_low_broken: List[Dict[str, Any]],
    neither_broken: List[Dict[str, Any]],
    both_broken: List[Dict[str, Any]],
    total_mondays: int,
    output_file: str
) -> None:
    """
    Create an Excel report with the analysis results.
    
    Args:
        only_high_broken: List of weeks where only high was broken
        only_low_broken: List of weeks where only low was broken
        neither_broken: List of weeks where neither was broken
        both_broken: List of weeks where both were broken
        total_mondays: Total number of Mondays analyzed
        output_file: Path to save the Excel file
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Partial Break Analysis"
        
        # Write summary
        ws['A1'] = "=== Monday Range Break Analysis ==="
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        ws['A3'] = f"Total number of Mondays analyzed: {total_mondays}"
        ws['A4'] = f"Number of weeks with only high broken: {len(only_high_broken)}"
        ws['A5'] = f"Number of weeks with only low broken: {len(only_low_broken)}"
        ws['A6'] = f"Number of weeks with neither broken: {len(neither_broken)}"
        ws['A7'] = f"Number of weeks with both broken: {len(both_broken)}"
        ws['A8'] = f"Percentage of weeks with incomplete breaks: {(len(only_high_broken) + len(only_low_broken) + len(neither_broken)) / total_mondays:.2%}"
        ws['A9'] = f"Percentage of weeks with both broken: {len(both_broken) / total_mondays:.2%}"
        
        # Write sections
        row = 11
        row = write_section(ws, "Weeks with Only High Broken:", only_high_broken, row)
        row = write_section(ws, "Weeks with Only Low Broken:", only_low_broken, row + 2)
        row = write_section(ws, "Weeks with Neither Level Broken:", neither_broken, row + 2)
        row = write_section(ws, "Weeks with Both Levels Broken:", both_broken, row + 2)
        
        # Adjust column widths
        for col, width in enumerate(['A', 'B', 'C', 'D', 'E', 'F', 'G'], 1):
            ws.column_dimensions[width].width = 15
        
        # Save the workbook
        wb.save(output_file)
        logger.info(f"Analysis complete! Results have been saved to '{output_file}'")
        
    except Exception as e:
        logger.error(f"Error creating Excel report: {str(e)}")
        raise

def main():
    try:
        # Load data
        df = load_data('filtered_candles.csv')
        
        # Analyze partial breaks
        only_high_broken, only_low_broken, neither_broken, both_broken, total_mondays = analyze_partial_breaks(df)
        
        # Create Excel report
        create_excel_report(
            only_high_broken, only_low_broken, neither_broken, both_broken,
            total_mondays, 'monday_partial_breaks.xlsx'
        )
        
    except Exception as e:
        logger.error(f"An error occurred during analysis: {str(e)}")
        raise

if __name__ == "__main__":
    main() 
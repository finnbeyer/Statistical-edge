import pandas as pd
import numpy as np
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import logging
from typing import Dict, List, Any, Optional
from datetime import datetime

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

def load_data(file_path: str) -> pd.DataFrame:
    """
    Load and prepare the data for analysis.
    
    Args:
        file_path: Path to the CSV file containing candle data
        
    Returns:
        DataFrame with prepared data
    """
    try:
        df = pd.read_csv(file_path, sep=';')
        df['Date'] = pd.to_datetime(df['Date'])
        df['DayOfWeek'] = df['Date'].dt.dayofweek
        df['Week'] = df['Date'].dt.isocalendar().week
        df['Year'] = df['Date'].dt.isocalendar().year
        return df
    except Exception as e:
        logger.error(f"Error loading data: {str(e)}")
        raise

def analyze_monday_ranges(df: pd.DataFrame) -> tuple[int, int, int, Dict[int, int], Dict[int, int], List[Dict[str, Any]]]:
    """
    Analyze Monday ranges and track when they are broken.
    
    Args:
        df: DataFrame containing the candle data
        
    Returns:
        Tuple containing:
        - total_mondays: Total number of Mondays analyzed
        - monday_highs_taken: Number of times Monday's high was broken
        - monday_lows_taken: Number of times Monday's low was broken
        - day_high_break: Counter of which days broke Monday's high
        - day_low_break: Counter of which days broke Monday's low
        - unbroken_weeks: List of weeks where neither high nor low was broken
    """
    monday_highs_taken = 0
    monday_lows_taken = 0
    total_mondays = 0
    day_high_break = Counter()
    day_low_break = Counter()
    unbroken_weeks = []

    for (week, year), week_data in df.groupby(['Week', 'Year']):
        monday_data = week_data[week_data['DayOfWeek'] == 0]
        if len(monday_data) == 0:
            continue
            
        total_mondays += 1
        monday_high = monday_data['High'].iloc[0]
        monday_low = monday_data['Low'].iloc[0]
        monday_date = monday_data['Date'].iloc[0]
        
        rest_of_week = week_data[week_data['DayOfWeek'].isin([1,2,3,4])]
        
        # Check high break
        high_break_days = rest_of_week[rest_of_week['High'] > monday_high]
        high_broken = not high_break_days.empty
        if high_broken:
            monday_highs_taken += 1
            first_break_day = high_break_days['DayOfWeek'].iloc[0]
            day_high_break[first_break_day] += 1
        
        # Check low break
        low_break_days = rest_of_week[rest_of_week['Low'] < monday_low]
        low_broken = not low_break_days.empty
        if low_broken:
            monday_lows_taken += 1
            first_break_day = low_break_days['DayOfWeek'].iloc[0]
            day_low_break[first_break_day] += 1
        
        # If neither high nor low was broken, store this week's info
        if not high_broken and not low_broken:
            unbroken_weeks.append({
                'Date': monday_date,
                'Week': week,
                'Year': year,
                'Monday High': monday_high,
                'Monday Low': monday_low
            })

    return total_mondays, monday_highs_taken, monday_lows_taken, day_high_break, day_low_break, unbroken_weeks

def calculate_probabilities(
    total_mondays: int,
    monday_highs_taken: int,
    monday_lows_taken: int,
    day_high_break: Dict[int, int],
    day_low_break: Dict[int, int]
) -> tuple[float, float, Dict[str, float], Dict[str, float]]:
    """
    Calculate various probabilities from the analysis results.
    
    Args:
        total_mondays: Total number of Mondays analyzed
        monday_highs_taken: Number of times Monday's high was broken
        monday_lows_taken: Number of times Monday's low was broken
        day_high_break: Counter of which days broke Monday's high
        day_low_break: Counter of which days broke Monday's low
        
    Returns:
        Tuple containing:
        - high_break_prob: Probability of Monday's high being broken
        - low_break_prob: Probability of Monday's low being broken
        - day_high_probs: Dictionary of day-specific probabilities for high breaks
        - day_low_probs: Dictionary of day-specific probabilities for low breaks
    """
    high_break_prob = monday_highs_taken / total_mondays if total_mondays > 0 else 0
    low_break_prob = monday_lows_taken / total_mondays if total_mondays > 0 else 0
    
    day_names = {1: 'Tuesday', 2: 'Wednesday', 3: 'Thursday', 4: 'Friday'}
    day_high_probs = {day_names[day]: count/monday_highs_taken if monday_highs_taken > 0 else 0 
                     for day, count in day_high_break.items()}
    day_low_probs = {day_names[day]: count/monday_lows_taken if monday_lows_taken > 0 else 0 
                    for day, count in day_low_break.items()}
    
    return high_break_prob, low_break_prob, day_high_probs, day_low_probs

def create_excel_report(
    total_mondays: int,
    monday_highs_taken: int,
    monday_lows_taken: int,
    high_break_prob: float,
    low_break_prob: float,
    day_high_probs: Dict[str, float],
    day_low_probs: Dict[str, float],
    day_high_break: Dict[int, int],
    day_low_break: Dict[int, int],
    unbroken_weeks: List[Dict[str, Any]],
    output_file: str
) -> None:
    """
    Create an Excel report with the analysis results.
    
    Args:
        total_mondays: Total number of Mondays analyzed
        monday_highs_taken: Number of times Monday's high was broken
        monday_lows_taken: Number of times Monday's low was broken
        high_break_prob: Probability of Monday's high being broken
        low_break_prob: Probability of Monday's low being broken
        day_high_probs: Dictionary of day-specific probabilities for high breaks
        day_low_probs: Dictionary of day-specific probabilities for low breaks
        day_high_break: Counter of which days broke Monday's high
        day_low_break: Counter of which days broke Monday's low
        unbroken_weeks: List of weeks where neither high nor low was broken
        output_file: Path to save the Excel file
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Analysis Results"
        
        # Style definitions
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell_alignment = Alignment(horizontal='left', vertical='center')
        
        # Write summary statistics with formatting
        ws['A1'] = "=== Monday Range Analysis ==="
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        ws['A3'] = f"Total number of Mondays analyzed: {total_mondays}"
        ws['A3'].font = header_font
        
        ws['A5'] = "High Break Analysis:"
        ws['A5'].font = header_font
        ws['A5'].fill = header_fill
        
        ws['A6'] = f"Number of times Monday's high was broken: {monday_highs_taken}"
        ws['A7'] = f"Probability of Monday's high being broken: {high_break_prob:.2%}"
        
        ws['A9'] = "Day-specific probabilities for high breaks:"
        ws['A9'].font = header_font
        ws['A9'].fill = header_fill
        
        row = 10
        for day, prob in day_high_probs.items():
            ws[f'A{row}'] = f"{day}: {prob:.2%}"
            row += 1
        
        row += 2
        ws[f'A{row}'] = "Low Break Analysis:"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        
        row += 1
        ws[f'A{row}'] = f"Number of times Monday's low was broken: {monday_lows_taken}"
        row += 1
        ws[f'A{row}'] = f"Probability of Monday's low being broken: {low_break_prob:.2%}"
        
        row += 2
        ws[f'A{row}'] = "Day-specific probabilities for low breaks:"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        
        row += 1
        for day, prob in day_low_probs.items():
            ws[f'A{row}'] = f"{day}: {prob:.2%}"
            row += 1
        
        # Add summary statistics
        row += 2
        ws[f'A{row}'] = "Summary Statistics:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
        
        row += 1
        ws[f'A{row}'] = f"Average days to break Monday's high: {sum(day * count for day, count in day_high_break.items()) / monday_highs_taken:.2f} (1=Tuesday, 2=Wednesday, etc.)"
        row += 1
        ws[f'A{row}'] = f"Average days to break Monday's low: {sum(day * count for day, count in day_low_break.items()) / monday_lows_taken:.2f} (1=Tuesday, 2=Wednesday, etc.)"
        
        # Add unbroken weeks analysis
        row += 2
        ws[f'A{row}'] = "Weeks Where Neither High Nor Low Was Broken:"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        
        row += 1
        ws[f'A{row}'] = f"Total number of unbroken weeks: {len(unbroken_weeks)}"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 2
        # Headers for unbroken weeks table
        ws[f'A{row}'] = "Date"
        ws[f'B{row}'] = "Week"
        ws[f'C{row}'] = "Year"
        ws[f'D{row}'] = "Monday High"
        ws[f'E{row}'] = "Monday Low"
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = header_fill
        
        # Add unbroken weeks data
        for week_data in unbroken_weeks:
            row += 1
            ws[f'A{row}'] = week_data['Date'].strftime('%Y-%m-%d')
            ws[f'B{row}'] = week_data['Week']
            ws[f'C{row}'] = week_data['Year']
            ws[f'D{row}'] = week_data['Monday High']
            ws[f'E{row}'] = week_data['Monday Low']
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        # Save the workbook
        wb.save(output_file)
        logger.info(f"Analysis complete! Results have been saved to '{output_file}'")
        
    except Exception as e:
        logger.error(f"Error creating Excel report: {str(e)}")
        raise

def main():
    try:
        # Load data
        df = load_data('filtered_candles.csv')
        
        # Analyze Monday ranges
        total_mondays, monday_highs_taken, monday_lows_taken, day_high_break, day_low_break, unbroken_weeks = analyze_monday_ranges(df)
        
        # Calculate probabilities
        high_break_prob, low_break_prob, day_high_probs, day_low_probs = calculate_probabilities(
            total_mondays, monday_highs_taken, monday_lows_taken, day_high_break, day_low_break
        )
        
        # Create Excel report
        create_excel_report(
            total_mondays, monday_highs_taken, monday_lows_taken,
            high_break_prob, low_break_prob, day_high_probs, day_low_probs,
            day_high_break, day_low_break, unbroken_weeks,
            'monday_analysis_results.xlsx'
        )
        
    except Exception as e:
        logger.error(f"An error occurred during analysis: {str(e)}")
        raise

if __name__ == "__main__":
    main() 
